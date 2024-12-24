import logging
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from collections import defaultdict

def normalize_name(name):
    """仅清理动态填充元素的名称"""
    name = name.lower().replace("_Fall2024", "").replace("-Fall2024", "")
    return ''.join(filter(str.isalnum, name))


def get_top_left_cell(worksheet, row, column):
    """获取合并单元格的左上角单元格"""
    cell = worksheet.cell(row=row, column=column)
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = merged_range.bounds
            return worksheet.cell(row=min_row, column=min_col)
    return cell

def safe_write_cell(worksheet, row, column, value):
    """安全写入单元格，避免合并单元格写入问题"""
    try:
        cell = get_top_left_cell(worksheet, row, column)
        cell.value = value
        logging.debug(f"Written value '{value}' at row {row}, column {column}")
    except ValueError as e:
        logging.error(f"Failed to write value '{value}' at row {row}, column {column}: {e}")

def save_to_excel(workbook, data, course_name):
    """保存成绩到 Excel"""
    course_mapping = {
        "OPS": {
            "columns": {
                "lab": "O",
                "quiz": "J",
                "midterm": "C",
                "final": "C"
            },
            "rows": {
                "lab_start_row": 63,
                "quiz_start_row": 63,
                "midterm_row": 66,
                "final_row": 67
            }
        },
        "CPR": {
            "columns": {
                "activity": "J",
                "quiz": "O",
                "ict": "D",
                "final": "D"
            },
            "rows": {
                "activity_start_row": 27,
                "quiz_start_row": 27,
                "ict_row": 30,
                "final_row": 31
            }
        },
        "APS": {
            "columns": {
                "test": "I",
                "ws": "O",
                "Vretta": "C",
                "BestPresentation": "C"
            },
            "rows": {
                "test_start_row": 14,
                "ws_start_row": 14,
                "Vretta_row": 20,
                "BestPresentation_row": 19
            }
        },
        "IPC": {
            "columns": {
                "ws": "J",
                "q": "O",
                "midterm": "C",
                "final": "C",
                "ms": "C"
            },
            "rows": {
                "ws_start_row": 52,
                "q_start_row": 52,
                "midterm_row": 58,
                "final_row": 59,
                "ms_row": 54
            }
        },
        "COM111": {
            "columns": {
                "Summary": "C",
                "Analyzing": "C",
                "Personal": "D",
                "Transfer": "C",
                "week2": "C",
                "Persuasive": "D",
                "Final": "C"
            },
            "rows": {
                "Summary_row": 41,
                "Analyzing_row": 42,
                "Personal_row": 42,
                "Transfer_start_row": 43,
                "week2_row": 45,
                "Persuasive_row": 45,
                "Final_row": 46
            }
        }
    }

    course_config = course_mapping.get(course_name, {})
    columns = course_config.get("columns", {})
    rows = course_config.get("rows", {})

    ws_averages = defaultdict(list)  # 处理 IPC 的 WS 平均分

    for item_name, grade in data:
        logging.debug(f"Processing item: {item_name}, grade: {grade}")
        matched = False
        normalized_item = normalize_name(item_name)

        for key, column_letter in columns.items():
            normalized_key = normalize_name(key)
            if normalized_key in normalized_item:
                matched = True
                item_number = 0
                if f'{key}_start_row' in rows:
                    item_number = int(''.join(filter(str.isdigit, item_name)) or '0')
                if course_name == "IPC" and key == "ms":
                    if "program" in normalized_item:
                        row, column = 56, column_index_from_string("C")
                        logging.debug(f"Matched 'MS3 - Program' for {item_name}, setting row {row}, column {column}")
                    elif "video" in normalized_item:
                        row, column = 56, column_index_from_string("D")
                        logging.debug(f"Matched 'MS3 - Video' for {item_name}, setting row {row}, column {column}")
                    else:
                        row = rows.get("ms_row", 1) + item_number - 1
                        column = column_index_from_string(column_letter)
                        logging.debug(f"Matched 'MS' for {item_name}, setting row {row}, column {column}")
                elif course_name == "APS" and key == "Vretta":
                    row, column = rows.get("Vretta_row", 1), column_index_from_string("C")
                    logging.debug(f"Matched key 'Vretta' for {item_name}, setting row {row}, column {column}")
                elif course_name == "APS" and key == "BestPresentation":
                    row, column = rows.get("BestPresentation_row", 1), column_index_from_string("C")
                    logging.debug(f"Matched key 'BestPresentation' for {item_name}, setting row {row}, column {column}")
                elif item_number == 0:
                    row = rows.get(f"{key}_row", 1)
                    logging.debug(f"Matched key '{key}' for {item_name}, setting row {row}")
                else:
                    row_key = f"{key}_start_row"
                    start_row = rows.get(row_key, 1)
                    row = start_row + item_number - 1
                    logging.debug(f"Matched key '{key}' for {item_name}, setting start row {start_row} and row {row}")

                column = column_index_from_string(column_letter)
                safe_write_cell(workbook["TOTAL"], row, column, grade)
                break

        if not matched:
            logging.warning(f"No matching key found for item: {item_name}")

    # 处理 IPC 的 WS 平均分
    if course_name == "IPC" and ws_averages:
        row_key = "ws_start_row"
        start_row = rows.get(row_key, 1)
        for ws_number, grades in ws_averages.items():
            average_grade = sum(grades) / len(grades)
            row = start_row + ws_number - 1
            column = column_index_from_string(columns.get("ws", "A"))
            safe_write_cell(workbook["TOTAL"], row, column, average_grade)

def main():
    template_path = "SEMESTER 1 MARKINGS.xlsx"
    output_path = "test1_Updated_Markings.xlsx"
    workbook = load_workbook(template_path)

    # OPS 数据
    ops_data = [
        ("Lab #2", 1.0), ("Lab #3", 1.0), ("Lab #4", 1.0), ("Lab #5", 1.0),
        ("Lab #8_fall2024", 1.0), ("Quiz 4", 1.0), ("Quiz 2", 0.8), ("Lab#6-fall2024", 1.0),
        ("Lab#7_Fall2024", 1.0), ("Lab#9-SharedLibraries", 1.0), ("Lab#10", 1.0),
        ("Quiz #1", 0.4), ("Lab #1", 1.0), ("Quiz 3", 1.0), ("Quiz 5-Fall2024", 1.0),
        ("Quiz 6", 0.935), ("Quiz 7-fall2024", 0.8), ("Midterm-Fall2024", 0.8967)
    ]
    save_to_excel(workbook, ops_data, "OPS")


    # CPR 数据
    cpr_data = [
        ("Activity 01", 0.9),  # 90 / 100
        ("Activity 02", 0.96),  # 96 / 100
        ("Activity 03", 0.96),  # 96 / 100
        ("Activity 04", 0.97),  # 97 / 100
        ("Activity 05", 0.98),  # 98 / 100
        ("Activity 06", 0.95),  # 95 / 100
        ("Activity 07", 0.95),  # 95 / 100
        ("Activity 08", 0.98),  # 98 / 100
        ("Activity 09", 0.95),  # 95 / 100
        ("Activity 10", 1.0),    # 100 / 100
        ("Quiz 01", 0.8),       # 80 / 100
        ("Quiz 02", 1.0),       # 100 / 100
        ("Quiz 03", 1.0),       # 100 / 100
        ("Quiz 04", 0.8),       # 80 / 100
        ("Quiz 05", 0.9),       # 90 / 100
        ("Quiz 06", 1.0),       # 100 / 100
        ("Quiz 07", 0.6),       # 60 / 100
        ("Quiz 08", 1.0),       # 100 / 100
        ("Quiz 09", 1.0),       # 100 / 100
        ("Quiz 10", 0.8),       # 80 / 100
        ("ICT News", 0.9),      # 90 / 100
    ]
    save_to_excel(workbook, cpr_data, "CPR")

    # APS 数据
    aps_data = [
        ("WS06", 0.9), ("WS07", 1.0), ("WS08", 1.0), ("WS09", 0.4), ("WS10", 1.06),
        ("PRES1", 0.76), ("PRES2", 0.0), ("PRES3", 0.0), ("Test 02", 0.7),
        ("Vretta Intromath", 0.97), ("WS01", 1.02), ("WS02", 1.06), ("WS03", 1.0),
        ("WS04", 1.0), ("WS05", 0.44), ("Test 01", 0.5667), ("BestPresentation", 0.76)
    ]
    save_to_excel(workbook, aps_data, "APS")

    # IPC 数据
    ipc_data = [
        ("Q2", 0.7), ("Q4", 1.0), ("Q6", 0.8), ("WS3-A", 1.0), ("WS3-B", 1.0),
        ("WS4-A", 1.0), ("WS4-B", 1.0), ("WS5-A", 1.0), ("WS5-B", 1.0),
        ("WS6", 1.0), ("Attendance", 1.0), ("MS1", 1.0), ("MS 2", 1.0),
        ("MS 3 - Program", 1.0), ("MS3 - Video", 1.0), ("Q1", 1.0),
        ("Q3", 0.6), ("Q5", 1.0), ("WS1", 1.0), ("WS2-A", 1.0), ("WS2-B", 1.0),
        ("Midterm Test", 0.86)
    ]
    save_to_excel(workbook, ipc_data, "IPC")

    # COM111 数据
    com111_data = [
        ("Transfer Assignment 2", 12.25 / 15),
        ("Week 2 assignment", 4.75 / 5),
        ("Personal Reflective Essay", 8.5 / 10),
        ("Summary test", 7 / 10),
        ("Persuasive Paragraph", 13 / 15),
        ("Analyzing Costabel's (2023) Article", 4.65 / 5),
        ("Transfer Assignment 1", 14.65 / 15)
    ]
    save_to_excel(workbook, com111_data, "COM111")

    workbook.save(output_path)
    logging.info(f"All grades saved to {output_path} successfully!")

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    main()
