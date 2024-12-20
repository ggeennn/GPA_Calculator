from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import logging
import os
import time
from openpyxl import load_workbook

# Functions: click_button, login_to_seneca, get_courses, get_course_grades
def click_element(driver, by, identifier, wait, element_name="element"):
    try:
        element = wait.until(EC.element_to_be_clickable((by, identifier)))
        if element:
            # 获取元素的边界框
            rect = driver.execute_script("return arguments[0].getBoundingClientRect();", element)
            # 检查元素是否在视口内
            if rect['top'] < 0 or rect['bottom'] > driver.execute_script("return window.innerHeight;"):
                # 如果元素不在视口内，滚动到元素
                time.sleep(1) # 也可以写在scroll命令后面？？？
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                # time.sleep(1)

        element.click()
        logging.info(f"Clicked {element_name} successfully!")
    except Exception as e:
        logging.error(f"Failed to click {element_name}: {e}")

def login_to_seneca(driver, wait, email, password):
    try:
        email_input = wait.until(EC.element_to_be_clickable((By.ID, "i0116")))
        email_input.send_keys(email)
        click_element(driver, By.ID, "idSIButton9", wait, "next")

        password_input = wait.until(EC.element_to_be_clickable((By.ID, "i0118")))
        password_input.send_keys(password)
        click_element(driver, By.ID, "idSIButton9", wait, "signin")
        click_element(driver, By.ID, "KmsiCheckboxField", wait, "checkbox")
        click_element(driver, By.ID, "idSIButton9", wait, "Yes")
        logging.info("Login successful!")
    except Exception as e:
        logging.error(f"Login failed: {e}")

r'''待优化
def get_courses(driver, wait):
    try:
        course_elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//bb-base-grades-student')))
        courses = []
        for course in course_elements:
            try:
                name = course.find_element(By.XPATH, './/h3//a[@class="bb-click-target"]').text.strip()
                link = course.find_element(By.XPATH, './/span[@ng-hide="baseGradesStudent.showAllGrades"]').get_attribute("href") # not working
                courses.append((name, link))
            except Exception as e:
                logging.error(f"Failed to fetch course details: {e}")
                continue
        return courses
    except Exception as e:
        logging.error(f"Error fetching courses: {e}")
        return []
'''

def get_course_grades(driver, course_name, wait):
    try:
        logging.info(f"Getting grades for course: {course_name}")
        # Get the parent item elements
        items = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@ng-repeat="grade in courseGradesStudent.userGrades"]')))
        data = []

        # Iterate over each item and extract Item Name and Grade
        for item in items:
            try:
                # Extract Item Name
                name_xpath= './/div[@class="name ellipsis student-grades__exception"]'
                item_name = item.find_element(By.XPATH, name_xpath).text.strip()
                # Extract Grade
                grades_xpath = './/div[@class="grade-color"]'
                raw_grade = item.find_element(By.XPATH, grades_xpath).text.strip()
                score, total = map(float, raw_grade.split('/'))
                grade = score / total
                data.append([item_name, grade])
            except Exception as e:
                logging.error(f"Error processing item: {e}")

        # Print the results
        print(f"\n\n{course_name} Grades Report:")
        for idx, (name, grade) in enumerate(data, start=1):
            print(f"{idx}. Item: {name}, Grade: {grade}")
        print("\n\n")
        save_to_excel(data, course_name)

        # Close the course view window
        click_element(driver, By.XPATH, '//*[@id="main-content"]/div[3]/div/div[3]/button', wait, "close")
    except Exception as e:
        logging.error(f"Failed to get grades for the course '{course_name}': {e}")

def save_to_excel(data, course_name, template_path="SEMESTER 1 MARKINGS.xlsx", output_path="Updated_Markings.xlsx"):
    # 学科对应的关键字与列位置
    course_mapping = {
        "COM111": {"columns": {"Test": "B", "Quiz": "C", "Assignment": "D"}},
        "IPC": {"columns": {"Test": "E", "Quiz": "F", "Assignment": "G"}},
        "OPS": {"columns": {"Test": "H", "Quiz": "I", "Assignment": "J"}},
        "CPR": {"columns": {"Test": "K", "Quiz": "L", "Assignment": "M"}},
        "APS": {"columns": {"Test": "N", "Quiz": "O", "Assignment": "P"}},
    }
    
    try:
        # 打开模板文件
        workbook = load_workbook(template_path)
        sheet = workbook["All Courses"]  # 假设所有数据填充到 "All Courses" 表

        # 获取当前课程对应的列映射
        course_config = course_mapping.get(course_name, {})
        columns = course_config.get("columns", {})
        
        # 查找填充起始行（动态定位空行）
        start_row = sheet.max_row + 1
        
        # 根据关键字分类填充数据
        for item_name, grade in data:
            column = None
            for key, col in columns.items():
                if key.lower() in item_name.lower():  # 匹配 item name 的关键字
                    column = col
                    break

            if column:
                sheet[f"A{start_row}"] = course_name  # A列存储课程名称
                sheet[f"{column}{start_row}"] = grade  # 填充到对应列
                start_row += 1

        # 保存新文件
        workbook.save(output_path)
        logging.info(f"Grades for {course_name} saved to {output_path} successfully!")
    except Exception as e:
        logging.error(f"Failed to save data for {course_name} to Excel: {e}")

def main():
    # Setup WebDriver
    options = Options()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    wait = WebDriverWait(driver, 20)
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    try:
        driver.get("https://learn.senecapolytechnic.ca/")
        logging.info("Browser launched successfully!")
  
        # Click "OK" button
        click_element(driver, By.ID, "agree_button", wait, "OK")
        # Click login button
        click_element(driver, By.ID, "bottom_Submit", wait, "Login")

        # Login
        email = os.getenv("SENECA_EMAIL")
        password = os.getenv("SENECA_PASSWORD")
        login_to_seneca(driver, wait, email, password)

        # 进入grades分页面
        time.sleep(3) #只能写在点击命令前面？？？
        click_element(driver, By.XPATH, '//bb-base-navigation-button[8]//a', wait, "Grades")
        
        # Wait for content to load
        try:
            time.sleep(1)
            main_container = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="main-content-inner"]')))
            logging.info("Main container located successfully.")
            scroll_height = driver.execute_script("return arguments[0].scrollHeight;", main_container)
            client_height = driver.execute_script("return arguments[0].clientHeight;", main_container)
            print(f"Scroll Height: {scroll_height}, Client Height: {client_height}")
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", main_container)
            # time.sleep(1) 也可以写在scroll命令前面？？？
        except Exception as e:
            logging.error(f"Failed to locate main container: {e}")

        # Fetch and process courses
        # courses = get_courses(driver, wait) 待优化
        courses = [
        ('IPC', '(//*[@id="card__733205_1"]/div/div/div[2]/div[2]/div/a)[2]'),
        ('OPS', '//*[@id="card__731947_1"]/div/div/div[2]/div[2]/div/a/span[1]'),
        ('CPR', '(//*[@id="card__737264_1"]/div/div/div[2]/div[2]/div/a/span[1])[2]'),
        ('APS', '(//*[@id="card__733550_1"]/div/div/div[2]/div[2]/div/a/span[1])[2]'),
        ('COM111', '(//*[@id="card__732569_1"]/div/div/div[2]/div[2]/div/a/span[1])[2]')
        ]
        for course_name, course_xpath in courses:
            logging.info(f"Fetching grades for {course_name}...")
            try:
                click_element(driver, By.XPATH, course_xpath, wait, course_name)
                get_course_grades(driver, course_name, wait)
            except Exception as e:
                logging.error(f"Select {course_name} failed: {e}")
            
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
