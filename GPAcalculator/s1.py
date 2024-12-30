import logging
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from collections import defaultdict
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import json
import os.path

def setup_logging():
    """Configure logging system"""
    # Remove all existing handlers
    logger = logging.getLogger()
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # Set root logger level
    logger.setLevel(logging.DEBUG)
    
    # Create formatter
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # File handler
    file_handler = logging.FileHandler('gpa_calculator.log', mode='w', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

def setup_webdriver():
    """Configure and initialize WebDriver"""
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), 
        options=options
    )
    return driver

def get_credentials():
    """Get login credentials"""
    email = os.getenv("SENECA_EMAIL")
    password = os.getenv("SENECA_PASSWORD")
    
    if not email or not password:
        raise ValueError("Missing SENECA_EMAIL or SENECA_PASSWORD environment variables")
    
    return email, password

def load_json_config(filename):
    config_path = os.path.join(os.path.dirname(__file__), filename)
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
            logging.info(f"Successfully loaded configuration from {filename}")
            return config
    except Exception as e:
        logging.error(f"Failed to load configuration from {filename}: {e}")
        return {}

def click_element(driver, by, identifier, wait, element_name="element"):
    try:
        element = wait.until(EC.element_to_be_clickable((by, identifier)))
        if element:
            # Get element boundary box
            rect = driver.execute_script("return arguments[0].getBoundingClientRect();", element)
            # Check if element is in view
            if rect['top'] < 0 or rect['bottom'] > driver.execute_script("return window.innerHeight;"):
                # If element is not in view, scroll
                time.sleep(1) # Can also be written after scroll command?
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                # time.sleep(1)

        element.click()
        logging.info(f"Clicked {element_name} successfully!")
    except Exception as e:
        logging.error(f"Failed to click {element_name}: {e}")

def login_to_seneca(driver, wait, email, password):
    try:
        email_input = wait.until(EC.element_to_be_clickable((By.ID, "i0116")))
        email_input.send_keys(email)
        click_element(driver, By.ID, "idSIButton9", wait, "next")

        password_input = wait.until(EC.element_to_be_clickable((By.ID, "i0118")))
        password_input.send_keys(password)
        click_element(driver, By.ID, "idSIButton9", wait, "signin")
        click_element(driver, By.ID, "KmsiCheckboxField", wait, "checkbox")
        click_element(driver, By.ID, "idSIButton9", wait, "Yes")
        logging.info("Login successful!")
    except Exception as e:
        logging.error(f"Login failed: {e}")

def get_course_grades(driver, course_name, wait, workbook, output_path):
    try:
        logging.info(f"Getting grades for course: {course_name}")
        # Get the parent item elements
        items = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@ng-repeat="grade in courseGradesStudent.userGrades"]')))
        data = []

        # Iterate over each item and extract Item Name and Grade
        for item in items:
            try:
                # Extract Item Name
                name_xpath= './/div[@class="name ellipsis student-grades__exception"]'
                item_name = item.find_element(By.XPATH, name_xpath).text.strip()
                # Extract Grade
                grades_xpath = './/div[@class="grade-color"]'
                raw_grade = item.find_element(By.XPATH, grades_xpath).text.strip()
                score, total = map(float, raw_grade.split('/'))
                grade = score / total
                data.append([item_name, grade])
            except Exception as e:
                logging.error(f"Error processing item: {e}")

        # Print the results
        print(f"\n{course_name} Grades Report:")
        for idx, (name, grade) in enumerate(data, start=1):
            print(f"{idx}. Item: {name}, Grade: {grade}")
        print("\n")

        save_to_excel(workbook, data, course_name)

        workbook.save(output_path)
        logging.info(f"All grades saved to {output_path} successfully!")

        # Close the course view window
        click_element(driver, By.XPATH, '//*[@id="main-content"]/div[3]/div/div[3]/button', wait, "close")
    except Exception as e:
        logging.error(f"Failed to get grades for the course '{course_name}': {e}")

def normalize_name(name):
    name = name.lower()
    name = name.replace("_fall2024", "").replace("-fall2024", "").replace("fall2024", "")
    return ''.join(filter(str.isalnum, name))

def get_top_left_cell(worksheet, row, column):
    """Get the top-left cell of a merged range"""
    cell = worksheet.cell(row=row, column=column)
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = merged_range.bounds
            return worksheet.cell(row=min_row, column=min_col)
    return cell

def safe_write_cell(worksheet, row, column, value):
    """Safely write to a cell, handling merged cells"""
    try:
        cell = get_top_left_cell(worksheet, row, column)
        cell.value = value
        logging.debug(f"Written value '{value}' at row {row}, column {column}")
    except ValueError as e:
        logging.error(f"Failed to write value '{value}' at row {row}, column {column}: {e}")

def save_to_excel(workbook, data, course_name):
    # Load course mapping from JSON file
    course_mapping = load_json_config('semester1_mapping.json')

    course_config = course_mapping.get(course_name, {})
    columns = course_config.get("columns", {})
    rows = course_config.get("rows", {})

    ws_averages = defaultdict(list)  # Workshop average

    # First loop: collect all WS scores
    for item_name, grade in data:
        logging.debug(f"Processing item: {item_name}, grade: {grade}")
        matched = False
        normalized_item = normalize_name(item_name)
        logging.debug(f"normalnizing item: {normalized_item}")

        # Process WS
        if course_name == "IPC" and "ws" in normalized_item :
            try:
                ws_number = int(''.join(filter(str.isdigit, normalized_item)))
                ws_averages[ws_number].append(grade)
                logging.debug(f"Added grade {grade} to Workshop {ws_number}")
                continue
            except ValueError:
                logging.warning(f"Could not extract workshop number from {item_name}")

        # Process other projects
        for key, column_letter in columns.items():
            normalized_key = normalize_name(key)
            if normalized_key in normalized_item:
                matched = True
                if course_name == "IPC" and key == "ms":
                    if "program" in normalized_item:
                        row = rows.get("ms_program_row", 56)  # From configuration file, default value is 56
                        column = column_index_from_string(columns.get("ms_program", "C"))  # From configuration file, default value is "C"
                        logging.debug(f"Matched 'MS3 - Program' for {item_name}, setting row {row}, column {column}")
                    elif "video" in normalized_item:
                        row = rows.get("ms_video_row", 56)  # From configuration file, default value is 56
                        column = column_index_from_string(columns.get("ms_video", "D"))  # From configuration file, default value is "D"
                        logging.debug(f"Matched 'MS3 - Video' for {item_name}, setting row {row}, column {column}")
                    else:
                        row = rows.get("ms_start_row", 54) + int(''.join(filter(str.isdigit, normalized_item)) or '0') - 1
                        column = column_index_from_string(column_letter)
                        logging.debug(f"Matched 'MS' for {item_name}, setting row {row}, column {column}")
                elif f'{key}_start_row' in rows:
                    item_number = int(''.join(filter(str.isdigit, normalized_item)) or '0')
                    row = rows.get(f"{key}_start_row") + (item_number - 1 if item_number > 0 else 0)
                    column = column_index_from_string(column_letter)
                else:
                    row = rows.get(f"{key}_row", 1)
                    column = column_index_from_string(column_letter)
                
                safe_write_cell(workbook["TOTAL"], row, column, grade)
                break

        if not matched:
            logging.warning(f"No matching key found for item: {item_name}")

    # Process WS averages
    if course_name == "IPC" and ws_averages:
        ws_start_row = rows.get("ws_start_row", 52)
        ws_column = column_index_from_string(columns.get("ws", "J"))
        
        for ws_number, grades in ws_averages.items():
            average_grade = sum(grades) / len(grades)
            row = ws_start_row + ws_number - 1
            safe_write_cell(workbook["TOTAL"], row, ws_column, average_grade)
            logging.info(f"Saved Workshop {ws_number} average grade: {average_grade} (from grades: {grades})")
            
def main():
    # Setup logging
    setup_logging()
    logging.info("Starting GPA calculator...")
    
    try:
        # Get login credentials
        email, password = get_credentials()
        
        # Initialize WebDriver
        driver = setup_webdriver()
        wait = WebDriverWait(driver, 20)
        
        template_path = "SEMESTER 1 MARKINGS.xlsx"
        output_path = "Updated_Markings.xlsx"
        
        # Load workbook
        try:
            workbook = load_workbook(template_path)
            logging.info(f"Successfully loaded template: {template_path}")
        except Exception as e:
            logging.error(f"Failed to load template: {e}")
            raise
        
        # Launch browser and login
        driver.get("https://learn.senecapolytechnic.ca/")
        logging.info("Browser launched successfully!")
        
        # Click initial buttons
        click_element(driver, By.ID, "agree_button", wait, "OK")
        click_element(driver, By.ID, "bottom_Submit", wait, "Login")
        
        # Login
        login_to_seneca(driver, wait, email, password)
        
        # Navigate to grades page
        time.sleep(3)
        click_element(driver, By.XPATH, '//a[@ui-sref="base.grades"]', wait, "Grades")
        
        # Wait for content to load
        try:
            time.sleep(1)
            main_container = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="main-content-inner"]')))
            logging.info("Main container located successfully.")
            scroll_height = driver.execute_script("return arguments[0].scrollHeight;", main_container)
            client_height = driver.execute_script("return arguments[0].clientHeight;", main_container)
            logging.debug(f"Scroll Height: {scroll_height}, Client Height: {client_height}")
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", main_container)
        except Exception as e:
            logging.error(f"Failed to locate main container: {e}")
            raise
        
        # Load course configuration from file
        course_config = load_json_config('course_config.json')
        if not course_config:
            raise ValueError("Failed to load course configuration")
        
        # Process each course
        for course in course_config['courses']:
            print(f"\n\n======================={course['name']}=======================")
            logging.info(f"Processing course: {course['name']}")
            try:
                click_element(driver, By.XPATH, course['xpath'], wait, course['name'])
                get_course_grades(driver, course['name'], wait, workbook, output_path)
            except Exception as e:
                logging.error(f"Failed to process {course['name']}: {e}")
                continue
                
    except Exception as e:
        logging.error(f"An error occurred in main execution: {e}")
    finally:
        driver.quit()
        print("\n")
        logging.info("Browser closed. Program finished.")

if __name__ == "__main__":
    main()
